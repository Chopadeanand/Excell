"""
sipl_dashboard.py — HiRATE Report → Interactive HTML Dashboard
==============================================================
Reads *_REPORT.xlsx files and builds HiRATE_Dashboard.html with:
  - Sidebar navigation per project
  - Slide 1: Category-wise chart (Total, Issues, % Issues) + 3 KPI cards
  - Slide 2: Division-wise chart sorted by % Issues descending

Charts rendered via Chart.js in browser — no matplotlib needed.
Dependencies: pip install openpyxl
"""

import os, sys, glob, tempfile, argparse, re, json
import openpyxl

DEFAULT_OUTPUT = "HiRATE_Dashboard.html"
HIRATE_LOGO_B64 = "/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCADhAOEDASIAAhEBAxEB/8QAHAABAAIDAQEBAAAAAAAAAAAAAAYHAwQFAggB/8QASxAAAQMCAgQICQgGCgMAAAAAAQACAwQFBhEHEiExExRBUWGBldIWIlRVVnGCkdMVIzJSYpOU0TNCU2NykhckNENkc6GisfCDo7L/xAAcAQEAAQUBAQAAAAAAAAAAAAAABwMEBQYIAQL/xAA8EQABAgQCBgYJBAEFAQAAAAABAAIDBAURITEGEkFRcaETFSJhgZEHFBZSU7HB0eEyktLwQhcjM2Ky8f/aAAwDAQACEQMRAD8AjqIi+lC6IiIiIiIiIiIiIiIiIiIiL9AJIABJJyAA2lfikWjqnp5MVwV1cCaC1RyXKr/y4G6+XW4Nb7S8JtiVUgwjFiNYNpso/Ix0cjo3tcx7SWua4ZEEbwV5Uj0htgmv7L3SM1KS+U0d0ib9QyjORnsyB46lHEBBFwvY8IwojmHYUREXqpIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIuzWS/I+iq5VgyFTfaxlthOeThBFlNOR0F3AtPWuMTkCeZbumOU0dxtGFG7BYreyOobycam+em9xc1vsK1m36kI9+CyEg2xdE3C3icPldbdklN50UNbtdU4brzGeilqtrT05SscOjhFx1n0NVUZxZJYKmQMpsQUklscXbmyvydA7LnErWD2isT2SRvdFKwskYS17SMi1w2EHrXkm/Wh23L2fbfUi7xY8R+LLyiIrtY5ERERERERERERERERERERERERERERERERERERERERERERERERd/R/R09Zi2idXbKCj166sdlmBDA0yvz6CG6vWq+vdyqLzeq671f9orqiSpk6HPcXEeoZ5Ke8KLRovvtzzLai7zxWemIO0RgiaoPqIbGw/xKtlip993hu5ZiE3o4DRvx+g+V/Fe4JpqeeOop5DHNE8SRvG9rgcwR6iArP0iGCqv8d+pGBlLfaWO5xtBz1HSD51vrEjZAqtVjWCX5Y0TuiObqnDdw2dFLVf8AOUzD94vmRfqxNXevXt6SC9m7EeGfIk+C46Iiy6wyIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiISAMzuCLuYEtsF1xbb6WrLRRMkNRWOdubBEDJJn7LSOsLxfcOGYjwxuZNloaXZuJGw4TbsNooBLVNI2iqqcpZAefVaYm9RUEXRxNd57/AIjuN7qM+Fr6mSoIP6oc4kN6hkOpc5YCI/XeXb1mYrgXnVyGA4DAckU00N1UQxh8h1UojpL/AE0lrkc7aGvkyMLsucStj95ULXuCWWCaOeCR0c0Tg+N7TkWuBzBHqK+WOLHBw2JCfqPDipdNFLBM+CdhjljcWSMO9rgciPevCkmkIxVl5gxDTNa2mv1JHcmBu5r3jKZvrErX+8KNrYQQRcLER4XQxHM3H++aIiL1UkREREREREREREREREREREREREREREREREREREREREXdpZjZ9HGJLznqz3Dg7LSn/M+cn/8AWwNz+2uETkMyt7StJ8n2/DmFWkB1HRcfrGjZ/WanJ+ThztiEQ6yraafqQj34K/p7e06J7o5nAfU+CgaIiwiuURERFYuHZvlfRRLTuOtUYcr9dvRS1Wwjqmbn/wCTpXIWTQ5WRR40ZaKuVsdFfaeS1TuduaZR8071iURnP1pUQy09RJT1DDHNE8xyMO9rgciPeFmZN+tDtuVCfbcMib8DxH4t5LGiIrtY5ERERERERERERERERERERb9JZb1WQNqKSz3KphdnqyQ0sj2nLYciBks3g5iP0evH4GXurcwhi664cE9NBI+e21Q1aqidM9jZAd5a5hDo38z2kEbN+WS/cSUeJJbfPiDBuLsR3K1xDXqqOSvl45bxyl7Q7x4+aRuznAyKoRoroYuG3CyctKy8Zlw4620YeY3jmOa0vBvEfo9ePwMvdTwbxH6PXj8DL3VFPCvFPpRfO0Zu8nhVin0nvnaM3eVr1gPdX16pL7zyUr8G8R+j14/Ay91PBvEfo9ePwMvdUU8KsU+k987Rm7yeFWKfSe+dozd5OsB7qeqS+88lK/BvEfo9ePwMvdTwbxH6PXj8DL3VFPCrFPpPfO0Zu8nhVin0nvnaM3eTrAe6nqkvvPJSvwcxH6PXj8DL3U8G8R+j14/Ay91RTwqxT6T3ztGbvJ4VYp9J752jN3k6wHup6pL7zyU8wvg+73DEVBSXG0XGloXzNNXNNTSRsjhb40hLiMh4od1qA4yvb8R4rul9e0t49UvlYw/qMJyY3qaGjqX5PiXElRBJBPiK8yxSNLJI318rmvaRkQQXZEEci5StpiZ6awtZVQ2HDZqQ74m5vy8sfNFtWi31V2u1Ha6FmvVVk7IIWncXvcGjPo2rVU70QRGhqrxi92z5EoiKU/4ufOKL3AyP6NQKhDZruDd6+oTQ5w1stvAYnkuHpDw6cKYxuFjbMaiCF4fTT7MpoXtDo3gjYc2kbRszBXAVg49iF3wBYb6wa1TaZHWesOeZ4M5y0zvVkZWZ/ZAVfL6jw+jiFqRdUnWbkcR47PDLwXqKSSGVk0LzHLG4PY4b2uBzB96uHFVku+Jq+DFVhslfXUd7pY657qSlfKyKdwLZ4yWgjWErXnLpCpxZoquriYGRVdRGwbmslcAOoFfcvH6EnC918nVeww35Z+I/+qxvAvGPopfez5e6ngXjH0UvvZ8vdVd8fr/L6v7935px+v8AL6v7935q76wHu81R9Vgd/L7KxPAvGPopfez5e6ngXjH0UvvZ8vdVd8fr/L6v7935px+v8vq/v3fmnWA93mnqsDv5fZWJ4F4x9FL72fL3U8C8Y+il97Pl7qrvj9f5fV/fu/NOP1/l9X9+7806wHu809Vgd/L7KxPAvGPopfez5e6ngXjH0UvvZ8vdVd8fr/L6v7935px+v8vq/v3fmnWA93mnqsDv5fZWJ4F4x9FL72fL3VjqcJYqpqeSoqcNXmCGNpc+SShka1oG8kkZALRwphivqbdFiDE94r7RYHn5lzZSamuI3tgYTtHPI7xR07l08RYmnuNBFZqCE22x07taGiZIXl7v2krztkf0nYOQBXUGK6IL6tgkaVl4LLuJuchhz3D57FwERFXWLRbVquFdarhDcLbVzUlXCdaOWJ2Tm/mDyg7DyrVRF6CWm4zXfrpcCXl3H7zh250tzkJ4x8kVccFPMf2nBvY7UceUNybnt5StXiWjLzXiztSn+CuUp7o7x3T2WIWq+2ykrLedjKkUcT6im9Ws3x2/ZO3LcdwVu+DDALgy53Db8gsrKzbo0UMivDQdpaD57fHFRfiWjLzZiztSn+CnEtGXmzFnalP8FXw6WJ9JFX0UVlrqCf8AQ1UFvgLH9B8TNrhytO0LFxkeQ2vs6DuLSJnTWmSsUwY0vEa4Zghv8lIEHQaejsESHHhlpyIb+FRnEtGXmzFnalP8FOJaMvNmLO1Kf4KvPjI8htfZ0HcTjI8htfZ0HcVv7f0b4T/Jv8lU/wBP6l8aH+38KjOJaMvNmLO1Kf4KcS0ZebMWdqU/wVefGR5Da+zoO4nGR5Da+zoO4nt/RvhP8m/yT2AqXxof7fwqM4loy82Ys7Up/gpxLRl5sxZ2pT/BV58ZHkNr7Og7icZHkNr7Og7ie39G+E/yb/JPYCpfGh/t/CoziWjLzZiztSn+Cs9yuVkiw1Bh/DdBXUtEKx1bUvrKhk0k0uoGM2ta0BrW62zL9Yq9bYW1dfDTuo7UxjnZvcbfAA1g2uP0OQAlUDi25RXrEtyudPBHTwVM7nQRxsDAyMbGDIbAdUDPpzWzUGsSlXY6NLQ3NDTa7gM+6xOW3itX0lo8zRWtZFiNcX7GtsbDvWXD1zt9LRXW1Xqkqau1XOBjJo6eZscjJI3h8cjXOBAI8YbtzivHEtGXmzFnalP8FaHGHR1vG4GMa5s3CxsIBaCHZgZHYQvoioraSpipbhQ2+1spK6mjqYWmggJaHDa36HI4OHUqtcqktSoAmJhhcL2wANuNyMFR0bpcxWHul4URrS0XGsL4X2ePzVDcS0ZebMWdqU/wU4loy82Ys7Up/gq8+MjyG19nQdxOMjyG19nQdxap7f0b4T/Jv8lt3sBUvjQ/2/hUZxLRl5sxZ2pT/BTiWjLzZiztSn+Crz4yPIbX2dB3E4yPIbX2dB3E9v6N8J/k3+SewFS+ND/b+FRnEtGXmzFnalP8FOJaMvNmLO1Kf4KvPjI8htfZ0HcTjI8htfZ0HcT2/o3wn+Tf5J7AVL40P9v4VGcS0ZebMWdqU/wU4loy82Ys7Up/gq8+MjyG19nQdxZqcmWOWd9LZoKaBuvPUTUEDIom87namz/kqtA03pcxEEKFAiOccgA2/wD6XxE0EqEJhe+PDAGZLfwqG4loy814s7Up/grPR/0f2ycVtDhq7V1VGM4YrrXRy0wdyOexkbS8D6pOR5VK9IeP6Kupn2bDlvomUpzbNcXUEcU0/IQwBo4NnSfGPRuVcLeIcGG5ocWWO42uPIkcyo/nJkwIpZCe19toaAPC63b1dbjeri+4XWrfVVLwAXuyAAG5rQNjWjkAAAWkiK5WJc4uJc43JRERF4iIiIiIiIu5hHFN3wxWPmt0rXwS7KikmGtDOPtN5xyOGRHOrfw9dbRiqmdPYnOiq2N16i2SvzmjHK5h/vGdI2jZmAqEWWlqKikqoqqknlp6iJwfHLE4tcxw3EEbisJWqBJ1iFqR29oZOGY/Hctl0f0onKK//bOtDObTl4bir7RRvCekCgvepRYokioLkdkdyADYJzzTAfQd9sbNu0DepVV009JMYaiMsflmOUOHIQdxHSFCFd0cnKNEtFF2HJwyP2PcfC6nei1+TrELpJd2O1pzH93rCiItfWaREREWjjG4my4Cula12rUVuVupjy5vGch6MowRnzuCoxx1WEjkCsTThcM7xQYeYfFtdPrTj9/MA93ubqD3qvCAQQdxXRejFN6upkKCR2iLniceWXgudNNKn1hVohaeyzsjwz5rexBQi136421ry9tJVSwNed7gx5aCfXkrR0T3H5RwPUW17iZ7PU67M/2Exz/0kDv5wqnuFXUV9fUV1XJwlRUSullfqgaznHMnIbBtPIpPoiurLZjikiqH6lLcmmgqDzCTINd1P1D1FXdap4qMhFltrhhxGI52Vpo3UxTqrDjjBt7HgcFaSL3NE+GZ8MgyfG4tcOYg5FeFzU5paSDmulgbi4REReIiL3BFLPM2GGN0kjzk1rRmSuLi7GNrwsX0dDxe7XxpLXDPXpqN3LrZfpHj6o2A7zsyWaotBm6xF1IDeyM3HIfnuWKq9alKTB6WZdbcNp4Bda71lrw/bmXK/wBQ6CKQZ09LHtqKr+AHc3nednrVSY1xjc8TyMhla2itsLs6eghceDZ9px3vf9o9OWS413uVfd7jLcbnWTVlXMc3yyuzJ5h0AcgGwci1FOFC0clKNDtCF3nNxzP2Hd53UEaRaWTdadqnsw9jR9d/yRERbAtVRERERERERERERERERERERTTBGPqyxwstV0ifc7KDshLspabPe6Fx3fwnxT0b1C0VKPAhx4ZhxWhzTmDkrmTnI8lFEaA4tcNoX0HTuo6+2tutnrGV9vccjI0ZOid9SRu9jvXsO8ZrwoPoKgqI7heruJJGU9NQ8AWteQ2SWU5MDhucA0PdkdxAKnCgbTCjStJnRDlibOF7HZjv++PFdD6J1mYq9PExHbY3tcbbbe5FtW408U76yt2UdFE+qqTln83GC4+/LLrWquJpOuHyXgB1M12VReagQgcogiIe8+ou1G+9Wei9N6xqcKER2QbngMeeXir3SCpCm06LMbQMOJwCqS73Cou11q7pVnOermfPJ0FxJy6s8lipWGSrhjAB15WtyPLmQFiX6ASQACSTkAN5XRS5kLiXaxxUh0l0FLbMf3ugoYGU9NBVFsUTdzG5A5D3qO7eQkHkI5CvczZWyubO2Rsg+kJAQ7rz2rwi+ozw+I5wFrk4bu5fQUFwF7stsv7drq+mBny5J2eJKP5m5+0vKiWha4casd3w+85yUzhcaYcursZMP/h3UVLVAGmlN9Rqr7Dsv7Q8c+d10dolU+saVCiE9oDVPEYIs0ULTTy1lTPFSUUAznqpnascY9fKTyAbSsKi2muGpqML2KuZK/itNNLSTQhx1BIfHY/V3ZluuM9/ihUNFaRAq0+IEd1hYnDbbZ3b/BVtJarFpVPfNQmaxFvC+3wXMxppFfNDLacK8NRULgWT1rhq1FUOUD9mw/VG08p3hV2NgyCIp+lZWDKQhBgNDWjIBc6VCozNQjGNMv1nH+4bkREVwrJERERERERERERERERERERERERERblkt093vNFaqbZNWTsgYeYucBn1Z59SL1rS4gDMq4cBUHyVo8tsbm6s9ykfXy579U+JF/taXe0umtq6ug466KkaGUtO1tPTtG5scYDG5dQC1VzjpJUOsKnGjA4XsOAwHna66gocgKfT4UuP8QL8dvNemNc97WMaXOcQGgcpKrLTNcm1mNJLdA/WprRE2hZkdhe3Myu9Ze5w9kK0GVsdooa6/Sta5lsp3Tsa7c+X6MTet5b7l8+SPklkfLK8ySPcXPcd7nE5knrUhejim6kCJOuGLjqjgM/M/JR16TKn/wAUi0/9j8h9V5W7YpYYL7bp6h+pDFVxPkd9VoeCT7lpLYt1JPcLhTUFKwPqKmZkMTScgXucGjbybSFJiiVhIcLZqQ6Wblb7xpFvFytdUKqjnkjdFKAQHZRMB37dhBHUosujiSy3HDt6ns91iZFWQBpkY14eBrNDhtGzcQuciqTLnvjPc8WJJuNxuu7gC8iwYwttzkP9XZLwdSOQwvGpJn7LifWArruFM6jrpqVxz4J5aDzjkPWMivncjMZFXthi4/LeCbRdHO1p4ozQVR5eEiADSTzujLD71H3pDpvTyLZpoxhnHgcPnbmpJ9GlT6OZiSTjg4XHEZ8vktpYrzQG8YSvdoa3WlkpuM04AzJlh8cAdJaHN61lWxbqp1FXwVbBmYZA/LnAO0dYUU0Wf6vn4UzsaRfhkeV1LVTk2z0pEl3ZOBC+dwQRmNxRdzHtnbYcZXS1xDKCKcvp8txhf48f+1wHUuGulgQRcLlqLDdCe6G7MGx8EREXqpoiIiIiIiIiIiIiIiIiIiIiIiIrA0H0OtiGtvj25stVI50Z/fy/Ns/0Lz7Kr9XTo6oPkzR3RlwynutQ+tfmMiI25xxD1HJ7vaWD0jqHV9MjRwcbWHE4DyvdbPofTvX6vCYR2W9o+H5suuiLNRU8lXVw0sX05Xhjes71zmxjojgxouTgujiQ0XKiOmO48RwtbbJG/Ka4ymtqADkREzNsQPQXa7vZCqdSTSXeY77jSvrKZ2tRxOFNSZHMcDGNVpHryLvaUbXS9JkG0+Shyzf8QB47T4m5XMmkNSNSqUWY2E2HAYBF0MN17LViK23SSJ0rKOriqHMaci4MeHZA9S566mFLSL7ia3WY1Bp+O1DYeFDNbUz5csxn71kViYQcYjQzO4txXQ0lYjhxZjGrvsFLJSxzsjaIpHAkarA3ePUo2pLpJwt4HYpksYrjXBkLJRMYuDz1gdmWZ5udRpeKpN9L07+m/Vc347ckVj6ELhr1V0w5I7ZWw8apgf20IJIA53Rl/wDKFXC38O3Sex3+gvFPmZKOoZMGg5awB2t6xmOtW05Ksm5d8CJk4EeauKTPup87CmW/4kHw28leKLaubIGVjn0jw+kma2emeNzongOYR1ELVXMszLvloz4MTNpIPEYLqOFEbFYHtyIuPFQzTdQcNSWTEDG7eDdbqg/aZ48Z62OI9hVkr3xLQC74IvdtDdaZkIrqcZZnhIc3EAc5YXhUQp/0SqHr1JhPJxaNU8W4cxY+K5906p3qVXeQOy/tDxz5oiItlWnIiIiIiIiIiIiIiIiIiIiIiIiLYttHNcrlS26mGc9XMyCP+J7g0f6lfQd0EEVUKOkGVJRxspacczI2ho9+WfWqx0I0HC4pnvUjc4rRSunaSMxwz/m4x73F3sqw1FnpJqFmwZNpz7R+Q+qmL0ZU7VgxZxw/UdUcBieaLbtk8EEkxnFQOEp5ImPgcGvjc5urrAkEZgE5bN+S1FmpKaernbBTROlkdua3/uwdKjKSjRoMwyJAF3ggjC+OzDapPjw2RIbmRP0kY7MFGho+wQBkJMQgD/EQ/DWpi3B+j7DNr4zcaq/trJW61LQiphM0o+sRwfiM+0d+3IFbeLcd2/DxfRYffBcruMw+tID6ald+7B2SPH1j4o2b9qqWuq6qvrJa2tqJampmdrSSyuLnPPOSVPOj8KsvZ01SiZ5NAbzIHIeO5QVpLN0GWJl6dADnbXXNhwxx+SwuyLiWggZ7ATnkPWtyx3Ors14pLrQuY2ppZBJEXt1m5jnHKtJFs60Rri0hwzC7GLsR3PFN4N2u7oXVRibETFHqDVbnls59q46Ii+oj3RHFzzclWTgzBGEMT2wz0eILsK6JutUUAp4jMwcrmguGuzPlG0bMwFv/ANGuFvPt7/CRd9VbR1NRR1cVXR1EtPUQu1o5YnFr2HnBG5WvhLH1vvupRYkfDbrmdjLgAG09Qf3oH6Nx+sPFPKAtdrzawxnS01wNs2kDHgfofPYtz0cjUGYcIFRhartjgTY8ccPkpRHDRUditlspamsqnUMToeGqI2sc6PWJYPFJ+jmR6sliWarpp6Scw1EZjeBnkdxHODuI6QsKgWpzUeamnxpgWeTiLWxGGXzU7ycvCloDIUH9IGGN8OK2bZU8TuMFSRm2N4Lh9Zv6w6xmFRuNLQbBiy52cfQpqhzYjzxnxmHraWlXSoRpvoNd1mxAwf2iA0VQQP7yH6JPSWOb/KpB9G9Q1YsWTccxrDiMDyt5KPvSVTulk4c20YsNjwP5VbIiKW1CaIiIiIiIiIiIiIiIiIiIiIizUdNPW1kFFSs156iVsMTed7iGtHvIRegEmwVwaMqD5N0exzvaBPd6p1QefgY82Rg+t3CHrC7S2rlDDSzR22kOdNQQso4ekRjVz6zmetc3E15tGE4A+8Z1Nwe3Whtcb9V5z3Old/dt6PpHkCgiqwJvSOtxRKtuAbX2ADC5O44nfuXRtPdK6O0eEJlwaAMe8nE23lbpZBT0EtzuVXHQW6E5SVMo2E/VYN73fZCrPG+kGoutPLaLFFJbbO/xZCT/AFirH7xw3N+wNnOSo9ivEl2xNcBV3SdpEYLYIIm6kMDfqsbyDp2k8pK46k7R7RSUo7Q/9UXa47O4bvmeSifSXTSZqxMGD2IW7aeP2RERbStJREREREREREREU1wRj6rskEdqu0L7pZgcmxa2U1LnvMLju/gPino3qz4eKV1tbdbRWR19vedXhmDJ0bvqSN3sd0HqXz2uphq/3bDlxFfaaowyEasjCNaOZn1HtOxzf+jIrWK/otKVhusezE2OH13jnuK3PRvTOapBEKJ24W7aOB+mSuxaWL6D5XwDeaFrdaala240+Z3GL9Jl643P9yYWxFZ8XNEdC1tvvGWb7dI/xZSN5gcd/PqHxhyZhdi2SikukT6iPxGv1JmOG9h8V4I9RKimWlZvRmrwnzLbAHPYWnA2PA5ZqYYseT0jpcRku7WDgR3g7LjYbr50RdPFNpksWJLjZpM86OpfE0n9ZgPiu625HrXMU+rm17Cxxa7MIiIi+UREREREREREREREREU90F2l9wxwK/g2vjtNNJWeOQG8IBqx5k7Bk463slQJbVPca6mt9VQU9VJFS1ZYaiNhyEurnqh3KQNY7N3uC+Ht1mloNrq6kozIEwyK9tw03tvtlzzVn4u0hUloL6HCk0dbX7Wy3UtzjiPKIAfpH94dnMNuaqqommqJ5KiolkmmkcXySSOLnPcd5JO0lY0VrIU+Xp8EQZduq0c+8naVdVWsTdVjGNMuudg2DgEREV6sWiIiIiIiIiIiIiIiIiIiIvTHOY9r2Oc17SHNc05EEbiDyFWdhHSNDWMZbsYveXgBsN2Y3N7RyCZo+mPtDxhy5qr0VpOyMvPQTBmGhzTsP9wPesjTKrNUyMI0s+x5HiFZWni1GK4WfEDHQyx3KjEb5oXh7JXxZNDw4bCHRmP3FVqto3GuNqFqdVSOoWzcO2BxzayTIjWbzZgnPLfsz3BaqrQIXQw2wwb2AFznhv71TqM02bmXx2t1dY3I7znbuuiIiqqyREREREREREREREREREREREREREREREREREREREREREREREREREREREREREREREREREREX//Z"
CUBE_LOGO_B64   = "/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCACzARkDASIAAhEBAxEB/8QAHQABAAIDAQEBAQAAAAAAAAAAAAUHAwQGCAIBCf/EAFgQAAEDAwIDBAQEDQ8KBwAAAAECAwQABREGEgchMRMiQVEIFDJhFXGBshYjMzY3QnJ0dZGhsbMXGCQ1OFJVYmeCkpSipdIlQ1Zjc3bC0dPkRFNUZJOVlv/EABoBAQADAQEBAAAAAAAAAAAAAAABAgMEBQb/xAAqEQEAAgIBAwMDAwUAAAAAAAAAAQIDEQQSITEFQVETcYEVIjIjYZHw8f/aAAwDAQACEQMRAD8A9l0pSgUpSgUpSgUpSgUpSgUpSgUpSgUpSgUpSgUpSgUpSgUpSgUpSgUpSgUpSgUpSgUpSgUpSgUpSgUpSgUpSgUpSgUpSgUpSgUpSgUpSgUpSgUpVb3TWd+1VdZWn+GLUVaYrhZuGpJrZcgw3AcKaZQCDJeHPIBCEH2lZ7tB3d6u1qslvcuN5uUO2w2hlyRLfS02ke9SiAK4lfGTRb274EF/1GE/b2WxS5jSvidQ2Wz8iq2LFwr0zFuLd6v4k6tvyeYud8UJC2znP0lvAaYGfBtCfeT1rr7tcrdZ4BmXGWzEjJKUb3FBKcqICUj3kkAComYiNyOIVxXhNlJe0PxCabJ9v6GZC8e8hAKvyVuWfixw/uc5u3fRE1bp7hw3DuzLlvfWfJKJCUKV8gNdSbkEJDkmJIjsn/Or2FA952qOB7zy86+rzabXere5b7xbYdxhujDkeUwl1tQ96VAg1FbxbwbbgIIyDkUqtHuHt40n+zOFl6+DW0HcvT1yWt+2PD963klcU+RbJQPFs1P6B1vE1OuVbZcCTY9RwAPhCzTCkvMZ5BaVJJS60r7VxBIPuOQLDrKUpQKUpQKUpQKUpQKUpQKUpQKUpQKUpQKUpQKUpQKUpQKUpQKUpQKUqD19qOPpHRd31LJaW8i3xVvJZR7Ty8YQ2n+MpRSke9QoOR1zPuOs9Wr4b6flvQ4EZpL2qLnHXtcYaWMtw2lDml51PNSuqG+fVaTXVvyLBonT0SDGYiW23RUoYjspw0yykqShOT4DctI8SSoda0eE2mZGl9GsMXNaHr7PcVcL1IT/AJ+a7hTpz4pTyQnyQhI8KxcSNOwNXW6fpO5SnWGLtGaQQ06Wy6ltwrW1uHMBaTg457SryqmS2ojvpEt7UV7Xb2/UpTMB6TJjuuNw0SwHnW0ABwpSoJKkp3p3EHluHKq54s2/6IdMpZ15BXNtG9BNugwVy3nlJWgnGwhaVADO5AztUvGQTnZ0qmxxNZqskG33Nu4aahtxW7rJ9YUhyKp8rcjMrcBLxylDZCleLZTkjAs7sIsWwsi5ONR0R2UFx5TgQG1JSBu38seWfKua+G05Ynqmde3t/vwrMTtSXCvjBpC22uz6UsmiNT2+zGIVwSphchAZBRlW7KipADoKnAShJ5KUCRmyY1yvUWXGaegJlQC4TFW25tKU7Vdnv3HGQgZPPkOZGcGpCHc9MT5DERGrIVwLo2sxky2D2g5cgEAFQ6cuYPjWXW8NyTapLLAbSuTDfjIUo7Uh1xICCo+Gdu3PvApnx2yatua6+3z+YLRMsUbWNtfkojszrXJeWcJZjy9y1dRhOUgKPI8sj8orU1/pNGqocO92Kam2altwL1ougQe4T1adHVbK+i0H4xhQBHnXSfDbWMbidEvC7RKhw0XAPvuymwnDW5RUSsg7Ep7NOEhZBz0BHe9XWJSV25LqPYdccdR9ypalD8hHKt4mYv07W90Nw01YNXadMqTCVbbvDfVDu1uWrcqHKRjejP2yTkKSr7ZCknxrp6rXVSE6O4u2TVTJ7K26oUix3dIGEmUAVQnz/GyFsZ8e0bHgKsqtUlKUoFKUoFKUoFKUoFKUoFKUoFKUoFKUoFKUoFKUoFKUoFKUoFV3xlIuF30HpZSh2V11K06+g/btxGnZmPi7RhqrEqv9eKaTxg4b9r9s5c0tf7T1XPzQug7qZJRGbSpSVrUtW1CEDKlq8h+InywCa4W7iDG1QH7nJtbKGY21qF24UprJ3buzOBnCU91JyQCRnNSXEiLepKbYmzuuNgvqbkhqR2Dq21D2G3MEIUSAc8uQKQQVA1SM+7atv8ePIat93dvFvtJeedhsR5sZ+9dj2SES0kdmwoISsLwEABScqb3JSrmy4a551b2lWY6vL0Q8bVb7a5eXVbo8VhUgvKWpzYgJJKk5Jx3c9K4a9aos9pnpdvkORftUhsSkWiIEOJtTRB2qWpaksxwBnLzqklZ3BJxtQJnWnraOFF1kesqguIYclF5pkvlhsOFwqS2RleEDknGTjGOeK8B8R9dXLWtzOn7AmTHsD039iQVO5duMhSgEyZa+rz6ztPeyEckpAxk649dMTEaTHh6S1j6R1seEq3l21XNhDS1zINmhm5pDQ8Vyn1MsDywlDoKiACfGa9HHizP1ZeHLFcLDco1hmsldpkXCS284tXfK0YShGGSlJ2HaUgoWgLPJKfKfEONC0zrK3aGnwrhCs9keiqvMV9banZEhSUKkvktFQJU2oJQApWxPIYJVXuLScG2RNHxWbWYwl3DUTbwdaIIfSl8PIKSPtPU207AOQaCUjAFaJdlfLdCVEENDBdW9kNsZKhgDqEqO1IGRzxgdOpAqP0ms2yC5Em9jJVa1+rPPtKUssd1KgCFJGB2am8lOR4nxr71lqm26RmouF5YmCE7FdJlNM7244ZbcecLhHsgoRy8yAOpFVZwo1GNJS9QJu6L7LXLcU+qIVGS+1KbSz2zBG4pW4G346ytBCVKLwAw2muP6FJvOSO0x2/6p0xvawuPsJybwa1OuNn1uDAVcohSMkPxsSGiPfvbTXYWea3crRDuLP1OUwh9HxKSFD89RV/QI3Du4olDYlq1OhwE52gNHIrV4PpkI4S6ORLz6wmwwQ7nru7BGfy111ncRK7qaUpUhSlKBSlKBSlKBSlKBSlKBSlKBSlKBSlKBSlKBSlKBSlKBVdcbT8GuaM1YSEtWTUsf1lR+1ZlJXDUc+ABkoUfck1YtROsrBB1VpS66buaSqHc4jkV7HUBaSNw94zke8Cgk3mmn2lNPNpcbUMKSoZBrWTbIgI3B91IOQh19a0/iUSD8vSuY4P6hm3vSxgX1SRqSxvG2XpA8ZDYGHR/EdQUOpPkvHUGuzqs0rbvMGn4tKVoUhaQpKhgpIyCPKvFfFb0ctUaO17F1Xw8hLutiauTU5ESOkKk2/Y4lZSlsn6chJHdCe9jCSDjcfU+qNdR7RfTYrdYb3qK5tMIlS49rabUYjKyoIW4pxaE5UUL2oBKztJCcc6ndM3u26jsUW9Wh8vw5SSUKUgoUCCUqSpKgClSVApUkgEEEHmKsPDPHjh9fNW8WZMzRU9eu7hNSlN2TCipYct8hsBsIkJyEs7kIGN5SSUOeIr0x6NHDLUeg9IxWtY3n4QnsoUiHDbVuZtrSsEtpXgFZznmeSNykowCoq5P0W/s98efw3G+fLq6NV6xt+n7hFtfqN0u10lNreag22MXnQ0ggKcUSQlCQSBlShknCcnlQT8lhmQ0Wn20rRnOD4EdCPI++sDdvjIcQs9s4Uc09q+twAjoQFEjPv61raV1BbNTWRq72l1xcdxS21JdaU0604hRQttxCgFIWlSSkpIBBFSlVmlZncwjTgvSClOM8INQQIxzMvDAs0NPip6WoR0AfK7n5DXbW2I1At0aCwMNR2ktIH8VIAH5BVeXVw604zQLQwO0s2jMXC4Lx3HLk6gpjs+RLbanHVeRWzVlVZJSlKBSlKBSlKBSlKBSlKBSlKBSlKBSlKBSlKBSlKBSlKBSlKBSlKCvOINku9l1I3xH0hDcm3FmOmNebU0QDdYaSVDYDy9YayooJ9oFSCeYx1+lNQWjVOn4l+sUxEy3y0bmnEgg5BIUlQPNKkkFJSQCCCCARUpVXX1lvhpr5OqIY7HS+p5zce+sJ9iJPc2tsTQOiQ4rY050yS2s9FEhOXzTuq4mqp2oNHXS0NLucZtmZFukZxxtLjQUG3kFtSTnCsKQeSglOCg5JrXgXf9XaW4y37gvqdVvubDcZ6/wbnGSW1qDz+9wLRzHeceWrGe4cgFQKSL/rz/bv3e1y/wByU/p26DS9GBYb478e1nOE3mMTj7uXWPgbqTiZxM1De+Ldnh6bg2uWhNnhW6c65vcYYWtaVqcQDtUlTq/tTv3kYSEpUf30Z/s48f8A8Lx/nTKkfQE/c8xfwlK+cKC3OHun5mnrLIbuk9qfdJ816fOeZa7JouuqyUtoJJCEpCUjJJITknJNRPErWU22zYmkNJMNXDWV1bK4rK8lmCwDtXMkEey0gnAHVxWEp8Smb4g6ni6P0lNv0lhyUpkJbjRWvqkp9aghplH8Za1JSPjz4VGcLNJytPWyRdL+8zM1ZelJk3uY3kpU7jCWW88wy0DsQPIEnmo0Elw/0rB0dppqzw3XZLhcXImTHzl6ZJcUVOvOHxUpRJ8gMAcgBU/SlApSlApSlApSlApSlApSlApSlApSoy96gsVkCTd7xBgFQylL76UKV8QJyfkoJOlc7btc6PuEhMeJqW2OPKOEoMhKVKPkAcZPxVPSn2IsdyTJebYZbSVOOOKCUoA6kk8gKDJStS2XS2XRC1224w5qWyAsx30uBJPTO0nFbZIAyeQoFK0Lbe7Nc3lM227wJrqU7lIjyUOKA6ZISTy5ivm436x26R6tcLzbob20K7N+UhtWD0OCc45GgkaVDp1XpdSglOpLMSegE5vn/aqWacbdbS40tLiFDKVJOQR7jQfVK0bneLRa1Npud0gwlOAlAkSENlQHXG4jPUVqfRZpX/Say/15r/FQTNKjIOobBOlIiwr5bJUhedjTMtC1qwMnAByeQJ+Ssi73ZkLUhd3gJUkkKBkoBBHUHnVbXrX+U6N6b9RerbFb9T6Xuenbq32kG5RXIr4HXatJBI8iM5B8CBWT4dsn8MW7+so/519sXi0vvJZYukF11ZwlCJCSpR9wB51WM2Oe0Wj/ACjcOa4L3uffOHdvXeF7rxAW7bLmT1VJjOKYcV/OLe8e5Qqsbd+72uX+5Kf07ddPxU0twHsTytTcQrBpqK5cpJSqZMi5LzxSVHJAPMhJPyVzmkdc+i3pG6KummbrpO0zVNFlT8VhSFlBIJTkJ6ZSPxVolF+jP9nHj/8AheP86ZUj6An7nmJ+EpXzhUrY+K/o62O53e52jVOnoc28uh25PNJWFSVgqIUs7eZytf4zXKm4+hzj6nof+pq/w0FnahSNTccLHY1gLt+l4JvklPUKlvFbEUH7lCZSvj2GrGrm9CaP0bpSG6vR1ht1pYnhDrhiMhAeAB2E+eAo4+M10aiEpKlEADqTQftKhn9VabYc7Ny9wQoHBw6FY/FUhb7hBuDZcgzGJKB1LTgVj48dKyrnxXnpraJn7oiYls0rVmXK3QlhEyfFjLI3BLryUEjz5npW1V4tEzMRKSlYZkuJDaDsyUzHbJ2hTrgQCfLJ8a+oz7EllL8d5t5pXsrbUFJPhyIp1Rvp33GSlKVYKUpQKUpQKUpQV5xw109pCxsxrYpIu08qSypQyGUD2nMeJ5gAHxOeeMGt+HHCqbrON9Eup7nMZYlqKm8HdIkD9+VKzhPlyJPuGM6npPrdPEBlK87EWxvZ/Tcz+WvRtqZjx7XEjxAlMdplCGgnoEBIAx8mKCpNScB7O5bnFWC5TGZiUkoRKUlxpw+RwkEZ8+ePI1LSNOXLS/Ay+Wu6Xdy5PiA8rnzQyNn1NBPMpHv+QAcqs6uZ4rfY11F+D3vmmgrz0Vf2nvv3y1801cz/ANRX9yapn0Vf2nvv3y18w1cz31Ff3J/NQedvRb+va5fg0/pEVg9IZhEni1EjLJCXYsZskdQC4sHH46z+i59e1y/Bp/SIrX9IkvJ4rxlRhl8Q45aGOqt68flxQdpK4A6cUyoRb3dm3cd1TvZLSD7wEpJ/GK5bgJcbpYuJUrSC5JehKVIacbCiW0uNE/TEjwztI9+RnoK2rhcuPa4jiHIcttBSQpTEdjfj3bcnPxc6/fR4maZt+on4E9icxqh8La7SX7JwcqbSOqVnGTu5nHXwoPr0rv2x0+f9TI/O3W9Z+BVknWiHNXfLkhUiOh1SUobwCpIOBy99aPpW4+EdP56djI/O3Xxa7pxyTbIqYNuUqKGUBghiOco2jb1Plig7TRXB+06W1NEvsa7zpD0XftbcSgJVuQpBzgZ6KqQmcNLdJmPyVXKWlTzinCAlOAVEny99RnDGdxSk6mU3rGGWbZ6sshRaaT9MynbzQc9N1WdXNyOJh5MRGWu9K2rFvKirZp9iXrpenlyHUspfeb7QAbsICiPd4VYNl4eQLXdo1xauElxcde8JUlODy+KuX099mZ378lfNXVuV4PovA4+St72rua3nX41pjipWdzr3Vb6SPCL9WPSlusX0Q/AfqU4S+29S9Y3/AEtSNu3ejHtZzk9OlUL+sY/lR/uD/uK9lryEKI645V/O17jR6UgdWG52oygKISTppnJHh/4evp3Qsf8AWMfyo/3B/wBxT9Yx/Kj/AHB/3FVt+rT6U/8A63UX/wCaZ/6FXL6IHEPjTqvidOtvEWRdXLSizuvNJlWhuKjtg6yE4UlpJJ2qXyz58uVB6ntcb1G1xYZc7T1dlDW/GN21IGceHSql1LerprPUAtFrKvUyspZbCsJWB1cX7vH3D31aWp1rb03c1tZ3piOlOPPYarjgm20b3OcVjtUxgEfEVDP5hXg+rWtmz4uJE6rbz9o9mOXvMVTcDhhaURwJs2W88R3i2UoSPiGCfxmsNu4ey7ZqViXb7w41DR3isD6b19j96QfM/iqw6V1/o/Djp1TWvje1vpU+FRcbv2/i/eX/ABqq3EewPiqpONmPohiZ6epjP9NVbiZnFIJAEZeP9mxXl4eZHG53I3S1tzH8Y38+WcW6b27JbjV9bEX79T8xdS3DL6xrb9yv9Iqq41lI1m7bWk6jZUiKHgUEobHfwcezz6Zqx+GX1jW37lf6RVacLPGf1O94rMft8TGp8wmlurJMukpSlfRtylKUClKUClKUFRekdo6TeLZH1FbWFPSbegtyG0DKlsE5yB47Tk48lE+FafCPi5aEWKLZNTyTEkRUBpmUpJU26gDCdxGdqgOWTyOM5ycVdNcPqjhToy/yFynbeuDJcOVuwl9mVHzKcFOffjNBj1LxZ0ZaIDj0a6tXSTtPZR4h371eGVeykeZJ+Q9K/NP6gj8UOHV1YjxXoL70dyI6lxJLaXFIOClfRQBIPmPEDlmPtvBDRUSQHXzc56Rz7ORIAT/YSk/lqxoEOLAhtQ4MZqNGaTtbaaQEpSPIAUHnTghq+Loe+3SyamS5CbfWlK1qQT2DyCQQoDngg9fDA8DkWfrTitpS2WGQ5bLrHuU9xspjsxzv7xHIqI5JA6nPPyqZ1pw+0vqxz1i6QSiZjb61HV2buPAE9FfzgcVztr4I6Khykvvm5XBKTkNSZA2fLsSkn4icUHL+i3YpSFXPUbyFJjuNiJHUR9UIVuWR7gQkZ88+VQfpAOIZ4vwXnFbUNxoylHyAcWSa9FxI8eJFaixWG2GGkhDbbaQlKEjoAB0Fcvqzh1pXVF2+FLxDfeldklrciQtA2gkjkDjxNBrP8V+H7KCtWomlY8EMOqJ+QJqmrXLVrjj7Hu1miutMGczJORgpaZCApSsdN2z8agKtlPBnQAUCbXJUPIzXcH+1XW6c07Y9OxlR7JbI8JC8FfZp7y8dNyjzV8poKY9K7ncdPj/USPzt12mn+K2gYtht8V+/FDzMVptxPqb5woJAIyEY610mstEad1c7FdvkV19UVKktFD628BWM+yRnoKgP1GNA/wAGyv667/ioJK1cT9D3S5R7dBvZelSXA20j1R5O5R6DJQAPlrsa4ez8KdF2m6xrnBgSESYzgdaUZbigFDpyJwa7igqPT32ZnfvyV81dW5UJF0tZo1+Ve2mHBNUtbhWXVEZWCDyzjxNTdeb6ZxMnFpet9d7TPb4nTPHWaxOyq/4scYND8L5Nvj6vnSoq7gha4/ZRVu7gggKztHL2hVgVVXHbgZpjjBMtUnUF0vEJVsbcbZEFbaQoLKSd29Cv3o6Yr0mjnf12nBX+G7j/APWu/wDKrg0bqO16t0vA1JZXXHbdPa7WOtxsoUpOSMlJ5jpXnf8AWT8Nv9JtWf8AzR/+jXoHh5paFonRVq0pbpEiREtjHYNOSCkuKTknvbQBnn4AUE46hLjSm1pCkqBSoHoQfCqXSJ+gNY71NKdY7yUk9H2SR0P74YHyjyq6q1bpboNzimNcIrcho/arHQ+YPUH3ivN9R4E8qK3pbpvXvEs706u8eURC1rpmVHDwujTORzQ93FD3YPX5M1FP8R7SL2zCjR35UZZ2KfQk53EgDajqofl8ga/XuGennHCpD1wZSeiEPJIH9JJP5amdP6Uslkc7aFF3P/8AnOq3rHxeA+TFYVj1S8xW3TWPeY3O/wAI/qSr7jdzv8U/+y/41VbiPYHxVC6g0tZr7JRIuTDjjiG+zSUuqT3ck+B99TYGBgVvxOJkw8nNltrV9a/G01rMWmflwvGr62Iv36n5i6luGX1jW37lf6RVSd/stvvkREW4trcaQ4HEhKynvAEeHxms9ot8W1W5qBDQpDDWdgUoqIySTzPvNKcTJXn25E66Zrr+/sRWevqbVKUr0mhSlKBSlKBSlKDjL1fNST9aStLaV+Coi7fBYmTp1xYcfT9OW4ltptpC0EkhlwqUVgJyjAVk4kNP6gnO3CHYr/ajBvTsR+S4GHUuxyhp1De5KshWF9olSQU5AyFYI54r9pN6VqT6I7JfpdkuTkVMOUpplt1qU0hSlN70LB7yC45tUkj21A7hgDSToN6M9aptu1VeGrnCakMPzZOyQuWiQ4hxzelSdqVb20lG0BKBlITt7tBFscSpUzUqY1p01drtEetXrbUaK2ymQFpkusrKluOoaCe4MJ3bjk4zg4km+JNsmMxlWKyXu9vOwkzX48Vtpt2I0pa0DtUvuN4XvadTsTuVltXLlz1bXw2eszkWTZdWXKJKjwDB7VbDTodQXlulS0qTgqy4cEYxjxBIOI8KIEUMOWe9S4UsQBBlSnYseSuSA6472xDjZSl7tH31bkgJJcO5CgEhITGoNcW+Fp7TmooUmM5aLzLjpEp3ISGHm1LSsdME4T186ik8Qo8/Vjttt01hyC1NgsIfjtB0uKeQ+paFkqASB2Se8ASM9Dnl00vTMV632KEmVJS3ZpDL7Klr7RbnZoUgBajzOQrmeua0rloqJO1Qu/OTXw4qREfLQSCn9jpeSB8R7ZWfiFBH2riZb7g/Z1J09qBi3XyUI1puTrLPq8wFtbiXU7XStCChskFxCSoEYB8ILR3Fx2TojTt31Bpm8IkXazGay/FZZ7CdIbjF9xhhBe7RKilDhR2gSlQQe90zF6a03qNq56PsCYuoWLZpy5KcSxL7D1aLFaYeaaCX0AKk+22lsEBQTkuDemu6s2gIFts2h7WibIdb0hgRlLSnMjEN2L3/AOa6VcvEDwoMVy4naahS7tH2zZPwZZFXpa47QWl9pKQpTbfe7zoStk7Tjk+3z73LFc+JsCCm4Sjp3UMi12+R6pIuTDTKmBI3hCmgO1DhKVnapYRsBB73I1p2/hHZoNjs9rZuc8C23NM1T3c3ymkpShMZzlzaDbbCPMhhGT1zs3fhuZse7W1jU9yh2W5zFTnYDbTRCHlOBxe1wp3bFOArKc9VK57TtoJJOvrOW4wMWemVIvb1lEQto7ZDzXaKWtQ3Y7Psm1PAgklspIGSE1m0JrKLq+KmXDtN1hxXYzUuNIkobUzJacztUhxpa057vNCiFjIJSARULatJrd41X7VTsaRGhIgMxoyS5hEiUtID8kJH2wabishXXuLAwOspovRadOXaXc3Ls9PkyY7UdxxUZplTwQVEOvdmkBx47sFZAGByAychFx9T3+23jUbmpLpZhZtPMoflKi2t4POoU0pfd+nLwQQBgJUVdAATWyriTbojcoXux3qySmmG32IkpDDjsxLjqGUBrsXXElRdcbb2qKSFLTkAEGpK6aMtl0TqZm4OPuxtRR0R5LaTtLaUtlGUKHMHnnPgRUMvhozLhTWLpf5ry30MCO5GjsRfVnGXkvtvhKEbVOhxDaiVgp7gASAVBQY71xBnxobCvoWvdrnJukOK/BnR23HHGn1KSktLZdW0okpI9vuYysJGDWdXE2C2Wor+m9QNXZy7/BBthbYU+h8xVykEqDpaKFNIyFBZAJAVtwrbtt6Mlvvol3nVFyuclM6NLTvQhtlsMbilCGkjancVKKl81K5DOEpCf2RoSG9q9GpFT5AeReG7qGglO3eiC5D2Z67drhV55HlQarfEmE7GbZZ07fXb4ue7AVZEoY9abcbQlxalK7XsUoDa217+0wQ4gDvKCajpHFKI3erc47HfgWsWq7SrsxKjkzIj8JyIjstqFKBP09fJO4LygoJBBMnI4etJvcy/W29zLfdn7o5cESENNrSgORmY7jBSoEKbUI7a/BW5IIIAxWrI4U2eYn/KN0uUt12FcY0x8rSh2Quatha3dyAAhSPVmwgJAAAHXFBsr4lW6EzK+iCw3ywSmYyJLESa2yt2YhTiGkhnsXXEqV2rjTewqCgpxGRhQNaF84lojIcgSrJqCwXRAhOqblRY72G35rcZOC2+W1ZUs5wslIySM7Uq+NWcP5U6z3K4XS53XU95agBm2hsx4jjJQ83ICmu4G+2U6wwrK+4S0gYSndmJt+kbvrDUMydfJN9ajIiwEIkzY7UdS3WJqZWxtlOdqQWkhS1c1Fw7ThIwHXNcRLY5ckNG13Vu1uXFVrbvC0tCKuUHSz2YHadrgugthZb2lWMKwQTs6G1tE1clD0KzXiLDfYMiJLktI7GQgK2nm2tXZr5g9m6ELwfZylQToW/h1FgXpEmJdpDNvRcXLiIaI7QUXXFqcUgvbe0LXaLK9mc9E7tg2Vs6X0O1ZNUO39d0fmSVRVRApTKG1uoUtKwp9SAO2cTtCUrVzAKupUokOVjcTbgrVKremZp+4O/Dkm2fAUTIuaGmlOAP83SCMISohSEJwrO4cs7+s+LEK16FVqCw2ifd33dNyNQMNBKEIZZbSgjtypaSg5c9kblfS3MAlOD2emLDGsLM1DDhdVLnyJqlqSAQXnFOFPLqAVYFcvG4VWJq2aytypk5TGqWHorp3J3Q47peUWmSQQEhyS+sZBwXMdABQdhbp8iVYm7jItsm3PrZLiokpTZcbPPkotrWj8Sj1rgzxGvKdB8PNQsaWkXWTqlEVUmPBcaT2JdiKfIR2zqATkcsnGAcnOM9pZ7Xc4zT7VyvztzS6yhtAVGbaDZAUFKG0cyrIyDyG3ljJrl7Xw6lQtIWPT30UyVDTxim0SBEbCmewaUyAsHIc3IUQenmMGgwas4lLiW+5vWWxXR+PCnJgG7KZbXCMgPpZcb2hztjtUVIK9mwKByrkasauBuPDb1qPcLYzqa5xbJOuJuSoDbbfcfU+H1gObd/Zqd3LKc9VEZ29yuztMeVFhdjNuDlwe7RxfbONoQdqlqUlGEADCUkJB6kJBOSSaDbpSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlApSlB/9k="

SCRIPT_DIR     = os.path.dirname(os.path.abspath(__file__))


# ══════════════════════════════════════════════════════════════════════════════
# DATA EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════

def extract_report_data(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['Sheet1']

    # Project name from second sheet name  e.g. ALL_MKTPL_Ratings_List → MKTPL
    data_sheet = next((s for s in wb.sheetnames if s != 'Sheet1'), None)
    proj_name  = "Project"
    if data_sheet:
        m = re.match(r'ALL_(.+?)_Ratings_List', data_sheet)
        proj_name = m.group(1) if m else data_sheet

    # Category data: rows 3-9, cols 10(label) 11(total) 12(issues) 13(pct)
    cat_labels, cat_total, cat_issues, cat_pct = [], [], [], []
    for r in range(3, 10):
        lbl = ws.cell(r, 10).value
        if lbl is None:
            continue
        if str(lbl).lower().startswith("overall"):
            continue
        cat_labels.append(str(lbl))
        cat_total.append(int(ws.cell(r, 11).value or 0))
        cat_issues.append(int(ws.cell(r, 12).value or 0))
        pv = ws.cell(r, 13).value
        cat_pct.append(round(float(pv) * 100, 2) if isinstance(pv, float) else 0.0)

    # KPI overall: row 3, cols 16(total) 17(satisfactory) 18(observations)
    total_kpi = int(ws.cell(3, 16).value or 0)
    sat_kpi   = int(ws.cell(3, 17).value or 0)
    obs_kpi   = int(ws.cell(3, 18).value or 0)
    sat_pct   = round(sat_kpi / total_kpi * 100, 1) if total_kpi else 0
    obs_pct   = round(obs_kpi / total_kpi * 100, 1) if total_kpi else 0

    # Division data: rows 201+, cols 1(label) 2(total) 3(issues) 4(pct)
    div_labels, div_total, div_issues, div_pct = [], [], [], []
    r = 201
    while ws.cell(r, 1).value:
        div_labels.append(str(ws.cell(r, 1).value))
        div_total.append(int(ws.cell(r, 2).value or 0))
        div_issues.append(int(ws.cell(r, 3).value or 0))
        pv = ws.cell(r, 4).value
        div_pct.append(round(float(pv) * 100, 2) if isinstance(pv, float) else 0.0)
        r += 1

    # Sort division by % Issues descending
    if div_labels:
        combined = sorted(zip(div_pct, div_labels, div_total, div_issues), reverse=True)
        div_pct, div_labels, div_total, div_issues = map(list, zip(*combined))

    return {
        "proj_name":    proj_name,
        "cat_labels":   cat_labels,
        "cat_total":    cat_total,
        "cat_issues":   cat_issues,
        "cat_pct":      cat_pct,
        "total":        total_kpi,
        "satisfactory": sat_kpi,
        "observations": obs_kpi,
        "sat_pct":      sat_pct,
        "obs_pct":      obs_pct,
        "div_labels":   div_labels,
        "div_total":    div_total,
        "div_issues":   div_issues,
        "div_pct":      div_pct,
    }


# ══════════════════════════════════════════════════════════════════════════════
# HTML BUILD  (Chart.js — no matplotlib)
# ══════════════════════════════════════════════════════════════════════════════

def build_html(projects):
    # Build JS PROJECTS object from all extracted data
    js_projects = {}
    for d in projects:
        js_projects[d["proj_name"]] = {
            "total":      d["total"],
            "sat":        d["satisfactory"],
            "obs":        d["observations"],
            "catLabels":  d["cat_labels"],
            "catTotal":   d["cat_total"],
            "catIssues":  d["cat_issues"],
            "catPct":     d["cat_pct"],
            "divLabels":  d["div_labels"],
            "divTotal":   d["div_total"],
            "divIssues":  d["div_issues"],
            "divPct":     d["div_pct"],
        }

    projects_json = json.dumps(js_projects, ensure_ascii=False)
    first_key     = projects[0]["proj_name"] if projects else ""
    hirate_logo   = HIRATE_LOGO_B64
    cube_logo     = CUBE_LOGO_B64

    hirate_src = f"data:image/png;base64,{HIRATE_LOGO_B64}"
    cube_src   = f"data:image/png;base64,{CUBE_LOGO_B64}"

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>HiRATE Audit Dashboard</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/chartjs-plugin-datalabels/2.2.0/chartjs-plugin-datalabels.min.js"></script>
<style>
:root {{
  --primary: #2c3e50; --accent: #3498db;
  --bg: #f0f2f5; --card: #fff;
  --sidebar: 240px; --radius: 10px;
  --shadow: 0 2px 12px rgba(0,0,0,0.08);
}}
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{
  font-family: 'Segoe UI', sans-serif;
  background: var(--bg); color: #333;
  height: 100vh; display: flex; overflow: hidden;
}}
aside {{
  width: var(--sidebar); background: var(--primary); color: #fff;
  display: flex; flex-direction: column; flex-shrink: 0;
  box-shadow: 3px 0 10px rgba(0,0,0,0.2);
}}
.brand {{
  padding: 22px 20px 18px;
  border-bottom: 1px solid rgba(255,255,255,0.1);
}}
.brand h1 {{ font-size: 24px; font-weight: 800; letter-spacing: 1px; }}
.brand p  {{ font-size: 10px; opacity: .5; text-transform: uppercase; letter-spacing: 2px; margin-top: 2px; }}
nav {{ flex: 1; overflow-y: auto; padding: 12px 0; }}
nav button {{
  width: 100%; background: none; border: none; color: rgba(255,255,255,.65);
  padding: 11px 20px; text-align: left; cursor: pointer; font-size: 13px;
  font-weight: 500; transition: all .2s; display: flex; align-items: center; gap: 10px;
}}
nav button:hover {{ background: rgba(255,255,255,.08); color: #fff; }}
nav button.active {{ background: var(--accent); color: #fff; font-weight: 600; }}
nav button .dot {{ width: 7px; height: 7px; border-radius: 50%; background: #2ecc71; flex-shrink: 0; }}
.nav-label {{ font-size: 9px; text-transform: uppercase; letter-spacing: 2px; opacity: .4; padding: 14px 20px 4px; }}
main {{ flex: 1; display: flex; flex-direction: column; overflow: hidden; }}
.tab-bar {{
  background: var(--primary); display: flex; padding: 0 24px; gap: 2px; flex-shrink: 0;
}}
.tab {{
  padding: 10px 22px; cursor: pointer; font-size: 13px; font-weight: 600;
  color: rgba(255,255,255,.5); border: none; background: none;
  border-radius: 8px 8px 0 0; border-bottom: 3px solid transparent; transition: all .2s;
}}
.tab.active {{ color: #333; background: var(--bg); border-bottom-color: var(--accent); }}
.tab:hover:not(.active) {{ color: rgba(255,255,255,.85); }}
.content {{ flex: 1; overflow-y: auto; padding: 20px 24px; display: flex; flex-direction: column; gap: 16px; }}
.page-hdr {{ display: flex; justify-content: space-between; align-items: center; }}
.page-hdr h2 {{ font-size: 18px; color: var(--primary); font-weight: 700; }}
.badge {{ background: #fff3cd; color: #856404; border: 1px solid #ffd96a; padding: 4px 12px; border-radius: 4px; font-size: 10px; font-weight: 700; letter-spacing: .5px; }}
.card {{ background: var(--card); border-radius: var(--radius); box-shadow: var(--shadow); padding: 16px 20px; }}
.card-title {{ font-size: 9px; font-weight: 700; color: #999; text-transform: uppercase; letter-spacing: 2px; margin-bottom: 12px; }}
.chart-wrap {{ position: relative; width: 100%; }}
.chart-wrap canvas {{ width: 100% !important; }}
.kpi-row {{ display: grid; grid-template-columns: repeat(3,1fr); gap: 14px; }}
.kpi {{
  background: var(--card); border-radius: var(--radius); box-shadow: var(--shadow);
  padding: 16px 18px; border-left: 5px solid #ddd;
}}
.kpi.blue  {{ border-left-color: #1F4E79; }}
.kpi.green {{ border-left-color: #27ae60; }}
.kpi.red   {{ border-left-color: #c0392b; }}
.kpi-lbl {{ font-size: 9px; font-weight: 700; color: #999; text-transform: uppercase; letter-spacing: 1.5px; margin-bottom: 8px; }}
.kpi-val {{ font-size: 32px; font-weight: 800; line-height: 1; }}
.kpi.blue  .kpi-val {{ color: #1F4E79; }}
.kpi.green .kpi-val {{ color: #27ae60; }}
.kpi.red   .kpi-val {{ color: #c0392b; }}
.kpi-sub {{ font-size: 11px; color: #999; margin-top: 4px; }}
.slide {{ display: none; flex-direction: column; gap: 16px; }}
.slide.active {{ display: flex; }}

.brand-logo {{ display: flex; align-items: center; gap: 10px; }}
.logo-hirate {{ width: 42px; height: 42px; object-fit: contain; flex-shrink: 0; }}
.topbar {{
  background: #fff; display: flex; align-items: center;
  justify-content: space-between; padding: 8px 24px;
  border-bottom: 1px solid #e0e0e0; flex-shrink: 0;
  box-shadow: 0 1px 4px rgba(0,0,0,0.06);
}}
.topbar-title {{ font-size: 13px; font-weight: 700; color: #2c3e50; letter-spacing: .3px; }}
.logo-cube {{ height: 36px; object-fit: contain; }}
</style>
</head>
<body>

<aside>
  <div class="brand">
    <div class="brand-logo">
      <img src="data:image/png;base64,__HIRATE_LOGO__" alt="HiRATE" class="logo-hirate"/>
      <div>
        <h1>HiRATE</h1>
        <p>Audit Report</p>
      </div>
    </div>
  </div>
  <div class="nav-label">Projects</div>
  <nav id="nav"></nav>
</aside>

<main>
  <div class="topbar">
    <span class="topbar-title">HiRATE Audit Dashboard</span>
    <img src="data:image/png;base64,__CUBE_LOGO__" alt="CubeTech" class="logo-cube"/>
  </div>
  <div class="tab-bar">
    <button class="tab active" onclick="switchTab(0)">&#x1F4CA; Category Overview</button>
    <button class="tab"        onclick="switchTab(1)">&#x1F4C8; Division Breakdown</button>
  </div>

  <div class="content">

    <div class="slide active" id="slide-0">
      <div class="page-hdr">
        <h2 id="title-0">&#x2014;</h2>
        <span class="badge">CONFIDENTIAL</span>
      </div>
      <div class="card">
        <div class="card-title">Category Wise Observations</div>
        <div class="chart-wrap" style="height:320px">
          <canvas id="catChart"></canvas>
        </div>
      </div>
      <div class="kpi-row">
        <div class="kpi blue">
          <div class="kpi-lbl">Total Observations</div>
          <div class="kpi-val" id="kpi-total">&#x2014;</div>
          <div class="kpi-sub">Audits Conducted</div>
        </div>
        <div class="kpi green">
          <div class="kpi-lbl">Satisfactory</div>
          <div class="kpi-val" id="kpi-sat">&#x2014;</div>
          <div class="kpi-sub" id="kpi-sat-n">&#x2014;</div>
        </div>
        <div class="kpi red">
          <div class="kpi-lbl">Issues Found</div>
          <div class="kpi-val" id="kpi-obs">&#x2014;</div>
          <div class="kpi-sub" id="kpi-obs-n">&#x2014;</div>
        </div>
      </div>
    </div>

    <div class="slide" id="slide-1">
      <div class="page-hdr">
        <h2 id="title-1">&#x2014;</h2>
        <span class="badge">CONFIDENTIAL</span>
      </div>
      <div class="card">
        <div class="card-title">Division Wise Observations &middot; Sorted by % of Issues &#x2193;</div>
        <div class="chart-wrap" style="height:420px">
          <canvas id="divChart"></canvas>
        </div>
      </div>
    </div>

  </div>
</main>

<script>
Chart.register(ChartDataLabels);

const PROJECTS = {projects_json};
const NAVY   = "#1F4E79";
const ORANGE = "#ED7D31";
const YELLOW = "#FFC000";

let catChart = null, divChart = null;

function makeChartConfig(labels, total, issues, pct, isDiv) {{
  const fs = isDiv ? 9 : 11;
  return {{
    type: 'bar',
    data: {{
      labels,
      datasets: [
        {{
          label: 'Total Audited',
          data: total,
          backgroundColor: NAVY,
          yAxisID: 'y', order: 2,
          barPercentage: 0.6, categoryPercentage: 0.75,
          datalabels: {{
            anchor: 'end', align: 'top', offset: 2,
            color: NAVY, font: {{ size: fs, weight: 'bold' }},
            formatter: v => v.toLocaleString()
          }}
        }},
        {{
          label: 'No of Issues',
          data: issues,
          backgroundColor: ORANGE,
          yAxisID: 'y', order: 2,
          barPercentage: 0.6, categoryPercentage: 0.75,
          datalabels: {{
            anchor: 'end', align: 'top', offset: 2,
            color: ORANGE, font: {{ size: fs, weight: 'bold' }},
            formatter: v => v > 0 ? v : ''
          }}
        }},
        {{
          label: '% of Issues',
          data: pct,
          type: 'line',
          yAxisID: 'y2', order: 1,
          borderColor: YELLOW, backgroundColor: YELLOW,
          pointStyle: 'rectRot', pointRadius: 6, pointHoverRadius: 8,
          borderWidth: 2, tension: 0,
          datalabels: {{
            anchor: 'end', align: 'top', offset: 4,
            color: '#a07000', font: {{ size: fs, weight: 'bold' }},
            formatter: v => v > 0 ? v.toFixed(1) + '%' : ''
          }}
        }}
      ]
    }},
    options: {{
      responsive: true, maintainAspectRatio: false,
      layout: {{ padding: {{ top: 28, bottom: isDiv ? 10 : 0 }} }},
      interaction: {{ mode: 'index', intersect: false }},
      plugins: {{
        legend: {{
          position: 'top', align: 'end',
          labels: {{ boxWidth: 12, font: {{ size: 11 }}, padding: 16 }}
        }},
        tooltip: {{
          callbacks: {{
            label: ctx => {{
              if (ctx.dataset.label === '% of Issues') return ' % of Issues: ' + ctx.parsed.y.toFixed(2) + '%';
              return ' ' + ctx.dataset.label + ': ' + ctx.parsed.y.toLocaleString();
            }}
          }}
        }},
        datalabels: {{}}
      }},
      scales: {{
        x: {{
          grid: {{ display: false }},
          ticks: {{
            font: {{ size: isDiv ? 9 : 11 }},
            maxRotation: isDiv ? 90 : 0,
            minRotation: isDiv ? 90 : 0,
            autoSkip: false,
          }},
          offset: true
        }},
        y: {{
          position: 'left', grid: {{ color: '#eee' }},
          ticks: {{ font: {{ size: 10 }} }}
        }},
        y2: {{
          position: 'right', grid: {{ drawOnChartArea: false }},
          ticks: {{ font: {{ size: 10 }}, color: YELLOW, callback: v => v.toFixed(1) + '%' }},
          title: {{ display: true, text: '% of Issues', color: YELLOW, font: {{ size: 11, weight: 'bold' }} }}
        }}
      }}
    }}
  }};
}}

const nav = document.getElementById('nav');
Object.keys(PROJECTS).forEach(key => {{
  const btn = document.createElement('button');
  btn.innerHTML = '<span class="dot"></span> ' + key;
  btn.onclick = () => show(key);
  btn.dataset.key = key;
  nav.appendChild(btn);
}});

function switchTab(i) {{
  document.querySelectorAll('.tab').forEach((t,j)  => t.classList.toggle('active', i===j));
  document.querySelectorAll('.slide').forEach((s,j) => s.classList.toggle('active', i===j));
}}

function animNum(el, target, decimals, suffix) {{
  decimals = decimals || 0; suffix = suffix || '';
  const dur = 700, start = Date.now();
  (function tick() {{
    const p = Math.min((Date.now()-start)/dur, 1);
    const v = target * p;
    el.textContent = decimals ? v.toFixed(decimals)+suffix : Math.round(v).toLocaleString()+suffix;
    if (p < 1) requestAnimationFrame(tick);
  }})();
}}

function show(key) {{
  const d = PROJECTS[key];
  nav.querySelectorAll('button').forEach(b => b.classList.toggle('active', b.dataset.key===key));
  const MONTHS = ['January','February','March','April','May','June','July','August','September','October','November','December'];
  const now = new Date();
  const curMonth = MONTHS[now.getMonth()];
  const curYear = now.getFullYear();
  const title = 'HiRATE Observations \u2014 ' + key + ' - ' + curMonth + ' ' + curYear;
  document.getElementById('title-0').textContent = title;
  document.getElementById('title-1').textContent = title;
  animNum(document.getElementById('kpi-total'), d.total);
  animNum(document.getElementById('kpi-sat'),   d.sat / d.total * 100, 1, '%');
  animNum(document.getElementById('kpi-obs'),   d.obs / d.total * 100, 1, '%');
  document.getElementById('kpi-sat-n').textContent = d.sat.toLocaleString() + ' observations';
  document.getElementById('kpi-obs-n').textContent = d.obs.toLocaleString() + ' issues';
  if (catChart) catChart.destroy();
  catChart = new Chart(document.getElementById('catChart'), makeChartConfig(d.catLabels, d.catTotal, d.catIssues, d.catPct, false));
  if (divChart) divChart.destroy();
  divChart = new Chart(document.getElementById('divChart'), makeChartConfig(d.divLabels, d.divTotal, d.divIssues, d.divPct, true));
}}

show({json.dumps(first_key)});
</script>
</body>
</html>""".replace("__HIRATE_LOGO__", HIRATE_LOGO_B64).replace("__CUBE_LOGO__", CUBE_LOGO_B64)


# ══════════════════════════════════════════════════════════════════════════════
# BUILD ENTRY POINTS
# ══════════════════════════════════════════════════════════════════════════════

def build_dashboard(report_files, output_path):
    projects = []
    for f in report_files:
        print(f"  Reading: {os.path.basename(f)}")
        try:
            d = extract_report_data(f)
            projects.append(d)
            print(f"    → {d['proj_name']}")
        except Exception as e:
            print(f"    ⚠ Skipped: {e}")

    if not projects:
        print("No valid report files found.")
        return False

    html = build_html(projects)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"\n✓  Saved: {output_path}  ({len(projects)} project(s))")
    return True


def find_report_files(folder):
    return sorted(glob.glob(os.path.join(folder, "*_REPORT.xlsx")))


def generate_dashboard_from_reports(report_bytes_list, output_path):
    """Called by sipl_app.py — report_bytes_list = [(filename, bytes), ...]"""
    tmp = []
    try:
        for fname, fbytes in report_bytes_list:
            tf = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False,
                 prefix=os.path.splitext(fname)[0] + "_")
            tf.write(fbytes)
            tf.close()
            tmp.append(tf.name)
        return build_dashboard(tmp, output_path)
    finally:
        for f in tmp:
            try:
                os.unlink(f)
            except:
                pass


def main():
    ap = argparse.ArgumentParser(description="HiRATE HTML Dashboard generator")
    ap.add_argument("files", nargs="*", help="*_REPORT.xlsx files")
    ap.add_argument("--output", "-o", default=None)
    args = ap.parse_args()

    files = [os.path.abspath(f) for f in args.files] if args.files \
            else find_report_files(SCRIPT_DIR)

    if not files:
        print(f"ERROR: No *_REPORT.xlsx files found in {SCRIPT_DIR}")
        sys.exit(1)

    out = os.path.abspath(args.output) if args.output \
          else os.path.join(SCRIPT_DIR, DEFAULT_OUTPUT)

    print(f"\nHiRATE Dashboard  ·  {len(files)} file(s)  →  {out}\n{'─'*50}")
    sys.exit(0 if build_dashboard(files, out) else 1)


if __name__ == "__main__":
    main()
