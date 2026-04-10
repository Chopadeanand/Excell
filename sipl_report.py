import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os

# Input file auto-detected from script folder (or pass as argument)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

def clean_df(df):
    """Pre-processing: remove rows where Rating in (1,5,10) and Remarks is 'NA', NaN, or blank.
    Called before both validation and report generation."""
    import math

    def is_na_remark(val):
        # Catches: float NaN, None, empty string, 'NA', 'na', 'N/A' etc.
        if val is None:
            return True
        if isinstance(val, float) and math.isnan(val):
            return True
        return str(val).strip().upper() in ('NA', 'N/A', '')

    na_mask = (
        df['HO Rating'].isin([1, 5, 10]) &
        df['HO Remarks'].apply(is_na_remark)
    )
    removed = int(na_mask.sum())
    if removed > 0:
        df = df[~na_mask].reset_index(drop=True)
    return df, removed

def validate_file(df):
    issues = []

    # Rule 1: Check for 'Median Opening' in Asset Type
    median_opening = df[df['Asset Type'].str.strip().str.lower() == 'median opening']
    if len(median_opening) > 0:
        issues.append(f"Rule 1 FAILED: Found {len(median_opening)} rows with Asset Type = 'Median Opening'")

    # Rule 2: Rating '-' with remarks other than '-'
    rating_dash = df[df['HO Rating'] == '-']
    bad_remarks_dash = rating_dash[rating_dash['HO Remarks'] != '-']
    if len(bad_remarks_dash) > 0:
        issues.append(f"Rule 2 FAILED: {len(bad_remarks_dash)} rows where Rating='-' but Remarks is not '-'")

    # Rule 3: Rating '10' with remarks other than '-'
    # Note: NA/NaN rows already removed by clean_df() before this is called
    import math
    def is_ok_remark_10(val):
        if val is None: return False
        if isinstance(val, float) and math.isnan(val): return False
        return str(val).strip() == '-'

    rating_10 = df[df['HO Rating'] == 10]
    bad_10 = rating_10[~rating_10['HO Remarks'].apply(is_ok_remark_10)]
    if len(bad_10) > 0:
        unique_remarks = bad_10['HO Remarks'].astype(str).unique().tolist()
        remarks_str = ', '.join(f"'{r}'" for r in unique_remarks[:5])
        issues.append(f"Rule 3 FAILED: {len(bad_10)} rows where Rating='10' but Remarks is not '-' (found: {remarks_str})")

    # Rule 4: Rating '1' or '5' with remarks '-' or NaN (NA already removed by clean_df)
    def is_bad_remark_1_5(val):
        if val is None: return True
        if isinstance(val, float) and math.isnan(val): return True
        return str(val).strip() == '-'

    rating_1_5 = df[df['HO Rating'].isin([1, 5])]
    bad_remarks_1_5 = rating_1_5[rating_1_5['HO Remarks'].apply(is_bad_remark_1_5)]
    if len(bad_remarks_1_5) > 0:
        issues.append(f"Rule 4 FAILED: {len(bad_remarks_1_5)} rows where Rating='1' or '5' but Remarks is '-'")

    return issues

def compute_stats(df, asset_type):
    rows = df[df['Asset Type'] == asset_type]
    audited = rows[rows['HO Rating'].astype(str).str.strip() != '-']
    issues = rows[rows['HO Rating'].astype(str).str.strip().isin(['1', '5'])]
    total = len(audited)
    issue_count = len(issues)
    pct = issue_count / total if total > 0 else None
    return total, issue_count, pct

def compute_category_stats(df, category):
    rows = df[df['Category'] == category]
    audited = rows[rows['HO Rating'].astype(str).str.strip() != '-']
    issues = rows[rows['HO Rating'].astype(str).str.strip().isin(['1', '5'])]
    total = len(audited)
    issue_count = len(issues)
    pct = issue_count / total if total > 0 else None
    return total, issue_count, pct

def build_report(df, output_file, report_title=None):
    from openpyxl.chart import BarChart, PieChart, Reference, Series

    # ── Pre-processing: remove all NA-remark rows for ratings 1,5,10 ──
    df, removed = clean_df(df)

    import re as _re_t
    # Step 1: project name from data
    _proj = None
    if 'Project Name' in df.columns:
        _v = df['Project Name'].dropna().astype(str).str.strip()
        _v = _v[_v != '']
        if not _v.empty:
            _proj = _v.iloc[0]
    # Step 2: month+year from filename
    _src = report_title if report_title else os.path.splitext(os.path.basename(output_file))[0]
    _months = ['january','february','march','april','may','june','july','august','september','october','november','december']
    _m = _re_t.search(r'(?<![a-z])(' + '|'.join(_months) + r')(?![a-z]).*?(\d{4})', _src.lower())
    _my = f"{_m.group(1).capitalize()} {_m.group(2)}" if _m else None
    # Step 3: build title
    if _proj and _my:   report_title = f"{_proj} - {_my}"
    elif _proj:         report_title = _proj
    elif _my:           report_title = _my
    else:
        _fb = _re_t.sub(r'[_-]+', ' ', _src).strip()
        report_title = ' '.join(_fb.split()).title()
    chart1_title = f"HiRATE Observations Category wise - {report_title}"
    chart3_title = f"HiRATE Observations Division wise - {report_title}"


    report_rows = [
        ('Road Signage and Furniture', 'MBCB-Semi Rigid Barrier'),
        ('Road Signage and Furniture', 'Pavement Markings'),
        ('Road Signage and Furniture', 'Signages'),
        ('Road Signage and Furniture', 'Delineators'),
        ('Road Signage and Furniture', 'Lightings'),
        ('Road Signage and Furniture', 'Kilometer Stones'),
        ('Road Signage and Furniture', 'Hectometer Stones'),
        ('Road Signage and Furniture', 'Traffic Blinkers and Signals'),
        ('Roadway', 'Shoulder'),
        ('Roadway', 'Drainage'),
        ('Roadway', 'Kerb'),
        ('Landscaping', 'Median'),
        ('Landscaping', 'Row'),
        ('Roadway', 'Embankment'),
        ('Roadway', 'Pavement'),
        ('Road Signage and Furniture', 'PGR-Pedestrain Guardrail (PGR)'),
        ('Structures', 'Wearing Coat On Deck Slab'),
        ('Structures', 'Drainage Spouts'),
        ('Structures', 'Rigid Crash Barriers'),
        ('Structures', 'Quadrant Pitching'),
        ('Structures', 'Structure Numbering'),
        ('Structures', 'Object Hazard Marker'),
        ('Structures', 'Approach Settlements'),
        ('Structures', 'Condition Of Clearance Of Vent'),
        ('Structures', 'Stagnation Of Rain Water'),
        ('Project Facilities', 'Bus Bay'),
        ('TMS', 'Traffic Lights'),
        ('TMS', 'Automatic Boom Barrier'),
        ('TMS', 'Operator Monitor'),
        ('TMS', 'User Fare Display (UFD)'),
        ('TMS', 'Overhead Lane Status Light (OHLS)'),
        ('TMS', 'Automatic Vehicle Classification and Counting system (AVCC)'),
        ('TMS', 'Weigh in Motion (WIM)'),
        ('TMS', 'Static Weigh Bridge (SWB)'),
        ('TMS', 'License Plate Indicatory Camera (LPIC)'),
        ('TMS', 'Operator Customized Keyboard'),
        ('TMS', 'Incident Camera'),
        ('ATMS', 'PTZ'),
        ('Project Facilities', 'Toilet Block'),
    ]

    summary_cats = [
        'Road Signage and Furniture', 'Roadway', 'Landscaping',
        'Structures', 'Project Facilities', 'ATMS',
    ]

    # ── Pre-compute all values from df directly ──
    def asset_stats(asset):
        rows = df[df['Asset Type'] == asset]
        total = len(rows[rows['HO Rating'].isin([10, 5, 1])])
        issues = len(rows[rows['HO Rating'].isin([5, 1])])
        pct = issues / total if total > 0 else None
        return total, issues, pct

    def cat_stats(cat):
        rows = df[df['Category'] == cat]
        total = len(rows[rows['HO Rating'].isin([10, 5, 1])])
        issues = len(rows[rows['HO Rating'].isin([5, 1])])
        pct = issues / total if total > 0 else None
        return total, issues, pct

    # Pre-compute category subtotals for F/G columns (each row shows its category total)
    cat_totals = {}
    for cat in df['Category'].unique():
        t, i, _ = cat_stats(cat)
        cat_totals[cat] = (t, i)

    # Summary stats
    summary_data = []
    overall_total, overall_issues = 0, 0
    for cat in summary_cats:
        t, i, p = cat_stats(cat)
        overall_total += t
        overall_issues += i
        # Skip rows where both Total Audited AND No of Issues are 0
        if t == 0 and i == 0:
            continue
        summary_data.append((cat, t, i, p))
    overall_pct = overall_issues / overall_total if overall_total > 0 else None

    # ── Build workbook ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    ws.column_dimensions['A'].width = 23
    ws.column_dimensions['B'].width = 50.664
    ws.column_dimensions['J'].width = 23
    ws.column_dimensions['K'].width = 13.555
    ws.column_dimensions['L'].width = 12.887
    ws.column_dimensions['R'].width = 11.555

    def pct_fmt(cell):
        cell.number_format = '0.00%'

    # ── Row 1: Headers ──
    ws['A1'] = 'Category'
    ws['C1'] = 'Total Audited'
    ws['D1'] = 'No of Issues'
    ws['E1'] = '% of Issues'

    # ── Row 2: Sub-headers ──
    ws['K2'] = 'Total Audited'
    ws['L2'] = 'No of Issues'
    ws['M2'] = '% of Issues'
    ws['P2'] = 'Total Audited'
    ws['Q2'] = 'Satisfactory'
    ws['R2'] = 'Observations'

    # ── Data rows 2-40: hardcoded computed values ──
    for idx, (cat, asset) in enumerate(report_rows):
        r = idx + 2
        total, issues, pct = asset_stats(asset)
        cat_total, cat_issues = cat_totals.get(cat, (0, 0))

        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=asset)
        ws.cell(row=r, column=3, value=total)
        ws.cell(row=r, column=4, value=issues)
        e = ws.cell(row=r, column=5, value=pct if pct is not None else '-')
        if pct is not None:
            pct_fmt(e)
        ws.cell(row=r, column=6, value=cat_total)
        ws.cell(row=r, column=7, value=cat_issues)

    # ── Summary table J3:M(3+n) — only non-zero rows ──
    for s_idx, (cat, t, i, p) in enumerate(summary_data):
        r = s_idx + 3
        ws.cell(row=r, column=10, value=cat)
        ws.cell(row=r, column=11, value=t)
        ws.cell(row=r, column=12, value=i)
        m = ws.cell(row=r, column=13, value=p if p is not None else '-')
        if p is not None:
            pct_fmt(m)

    # Overall Audited row — immediately after last non-zero category row
    overall_row = 3 + len(summary_data)
    ws.cell(row=overall_row, column=10, value='Overall Audited')
    ws.cell(row=overall_row, column=11, value=overall_total)
    ws.cell(row=overall_row, column=12, value=overall_issues)
    m9 = ws.cell(row=overall_row, column=13, value=overall_pct if overall_pct is not None else '-')
    if overall_pct is not None:
        pct_fmt(m9)

    # ── Overall table P3:S3 ──
    satisfactory = overall_total - overall_issues
    ws['P3'] = overall_total
    ws['Q3'] = satisfactory
    ws['R3'] = overall_issues
    s3 = ws['S3']
    s3.value = overall_pct if overall_pct is not None else '-'
    if overall_pct is not None:
        pct_fmt(s3)

    from openpyxl.chart.series import DataPoint
    from openpyxl.drawing.fill import PatternFillProperties
    from openpyxl.chart.data_source import NumDataSource, NumRef
    from openpyxl.chart.label import DataLabel, DataLabelList
    from openpyxl.chart.marker import Marker
    from copy import deepcopy

    # (solid_fill helper not needed - using graphicalProperties directly)

    # ══════════════════════════════════════════════════
    # CHARTS - Direct XML for pixel-perfect styling
    # ══════════════════════════════════════════════════
    from openpyxl.chart.series import DataPoint
    import xml.etree.ElementTree as etree

    nsC = "http://schemas.openxmlformats.org/drawingml/2006/chart"
    nsA = "http://schemas.openxmlformats.org/drawingml/2006/main"

    def make_spPr(color, border_color=None, border_width=9525):
        """Solid fill spPr element"""
        sp = etree.Element("{%s}spPr" % nsC)
        sf = etree.SubElement(sp, "{%s}solidFill" % nsA)
        etree.SubElement(sf, "{%s}srgbClr" % nsA, val=color)
        ln = etree.SubElement(sp, "{%s}ln" % nsA, w=str(border_width))
        if border_color:
            sf2 = etree.SubElement(ln, "{%s}solidFill" % nsA)
            etree.SubElement(sf2, "{%s}srgbClr" % nsA, val=border_color)
        else:
            etree.SubElement(ln, "{%s}solidFill" % nsA)
            # white border between slices
            sf2 = etree.SubElement(ln, "{%s}solidFill" % nsA)
            etree.SubElement(sf2, "{%s}srgbClr" % nsA, val="FFFFFF")
        return sp

    def styled_series(chart, ref, color, show_val=True, title_from_data=True):
        """Append a bar series with solid color and data labels"""
        s = Series(ref, title_from_data=title_from_data)
        s.graphicalProperties.solidFill = color
        s.graphicalProperties.line.solidFill = color
        s.dLbls = DataLabelList()
        s.dLbls.showVal = show_val
        s.dLbls.showLegendKey = False
        s.dLbls.showCatName = False
        s.dLbls.showSerName = False
        s.dLbls.showPercent = False
        chart.append(s)

    # ─────────────────────────────────────────────────
    # CHART 1: Clustered Bar — Category wise
    # Matches: dark navy bars, orange bars, green bars
    # With data table below showing exact values
    # ─────────────────────────────────────────────────
    chart1 = BarChart()
    chart1.type = "col"
    chart1.grouping = "clustered"
    chart1.style = 2
    chart1.width = 18
    chart1.height = 12
    chart1.gapWidth = 150
    chart1.overlap = 0

    cats1 = Reference(ws, min_col=10, min_row=3, max_row=overall_row)
    styled_series(chart1, Reference(ws, min_col=11, min_row=2, max_row=overall_row), "1F4E79")  # Total Audited
    styled_series(chart1, Reference(ws, min_col=12, min_row=2, max_row=overall_row), "ED7D31")  # No of Issues
    styled_series(chart1, Reference(ws, min_col=13, min_row=2, max_row=overall_row), "92D050")  # % of Issues
    chart1.title = chart1_title
    chart1.set_categories(cats1)
    chart1.legend = None
    chart1.y_axis.majorGridlines = None  # remove gridlines

    # Add data table below the chart (like reference image)
    ws.add_chart(chart1, "I11")

    # ─────────────────────────────────────────────────
    # CHART 2: Pie — Observations Summary (PREMIUM)
    # Lime green 94% + exploded orange slice 6%
    # Value AND % shown on each slice
    # ─────────────────────────────────────────────────
    chart2 = PieChart()
    chart2.title = "OBSERVATIONS SUMMARY"
    chart2.style = 2
    chart2.width = 11
    chart2.height = 12

    cats2 = Reference(ws, min_col=17, min_row=2, max_row=2, max_col=18)
    pie_data = Reference(ws, min_col=17, min_row=3, max_col=18, max_row=3)
    series_pie = Series(pie_data)

    # Slice 0: Satisfactory — same lime green as reference
    pt0 = DataPoint(idx=0)
    pt0.graphicalProperties.solidFill = "92D050"
    pt0.graphicalProperties.line.solidFill = "FFFFFF"

    # Slice 1: Observations — orange, exploded
    pt1 = DataPoint(idx=1)
    pt1.graphicalProperties.solidFill = "ED7D31"
    pt1.graphicalProperties.line.solidFill = "FFFFFF"
    pt1.explosion = 8

    series_pie.dPt = [pt0, pt1]

    # Show BOTH value and % on slices (like reference: "3250\n94%" and "221\n6%")
    series_pie.dLbls = DataLabelList()
    series_pie.dLbls.showVal = True
    series_pie.dLbls.showPercent = True
    series_pie.dLbls.showCatName = False
    series_pie.dLbls.showSerName = False
    series_pie.dLbls.showLegendKey = False
    series_pie.dLbls.separator = "\n"

    chart2.append(series_pie)
    chart2.set_categories(cats2)
    chart2.legend.position = "r"
    ws.add_chart(chart2, "R10")

    # ─────────────────────────────────────────────────
    # CHART 3: Division wise — sorted by Issues DESC, zeros removed
    # Data written to row 200+ area (below visible content, avoids hidden-col issue)
    # ─────────────────────────────────────────────────

    # Build sorted, filtered data for chart3
    chart3_data = []
    for cat, asset in report_rows:
        total, issues, pct = asset_stats(asset)
        if issues > 0:  # exclude zero-issue rows
            chart3_data.append((asset, total, issues, pct))
    # Sort by issues descending
    chart3_data.sort(key=lambda x: x[2], reverse=True)

    # Write staging data at row 200+ in col A-D (visible but far below chart area)
    STAGE_ROW = 200
    SC = 1  # staging start column
    ws.cell(row=STAGE_ROW, column=SC,   value='Asset Type')
    ws.cell(row=STAGE_ROW, column=SC+1, value='Total Audited')
    ws.cell(row=STAGE_ROW, column=SC+2, value='No of Issues')
    ws.cell(row=STAGE_ROW, column=SC+3, value='% of Issues')
    for i3, (asset, total, issues, pct) in enumerate(chart3_data):
        rr = STAGE_ROW + i3 + 1
        ws.cell(row=rr, column=SC,   value=asset)
        ws.cell(row=rr, column=SC+1, value=total)
        ws.cell(row=rr, column=SC+2, value=issues)
        e3 = ws.cell(row=rr, column=SC+3, value=pct if pct is not None else '-')
        if pct is not None:
            pct_fmt(e3)

    n3 = len(chart3_data)
    max_r3 = STAGE_ROW + n3

    chart3 = BarChart()
    chart3.type = "col"
    chart3.grouping = "clustered"
    chart3.style = 2
    chart3.width = max(16, n3 * 1.2)
    chart3.height = 12
    chart3.gapWidth = 100

    cats3 = Reference(ws, min_col=SC,   min_row=STAGE_ROW+1, max_row=max_r3)
    styled_series(chart3, Reference(ws, min_col=SC+1, min_row=STAGE_ROW, max_row=max_r3), "1F4E79")
    styled_series(chart3, Reference(ws, min_col=SC+2, min_row=STAGE_ROW, max_row=max_r3), "ED7D31")
    styled_series(chart3, Reference(ws, min_col=SC+3, min_row=STAGE_ROW, max_row=max_r3), "92D050")
    chart3.title = chart3_title
    chart3.set_categories(cats3)
    chart3.legend = None
    chart3.y_axis.majorGridlines = None

    ws.add_chart(chart3, "A41")
    # ── Attach cleaned input data as 'ALL_SIPL_Ratings_List' sheet ──
    # Sheet name from project name in data
    _sn_proj = _proj if _proj else 'SIPL'
    _sheet_name = ('ALL_' + _sn_proj + '_Ratings_List')[:31]
    ws_data = wb.create_sheet(_sheet_name)
    # Write header
    for col_idx, col_name in enumerate(df.columns, 1):
        ws_data.cell(row=1, column=col_idx, value=col_name)
    # Write data rows
    import math
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            if isinstance(value, float) and math.isnan(value):
                value = None
            ws_data.cell(row=row_idx, column=col_idx, value=value)

    output_dir = os.path.dirname(output_file)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    wb.save(output_file)

    # ── Post-process: inject data table XML into bar charts ──
    import zipfile as _zf, io as _io, xml.etree.ElementTree as _et, re as _re
    _nsC = "http://schemas.openxmlformats.org/drawingml/2006/chart"
    def _inject_dt(xml_bytes):
        import xml.etree.ElementTree as ET
        import re
        # Register namespaces to preserve them on output
        for pfx, uri in re.findall(r'xmlns:([A-Za-z0-9_]+)="([^"]+)"', xml_bytes.decode('utf-8','ignore')):
            try: ET.register_namespace(pfx, uri)
            except Exception: pass
        ET.register_namespace('', 'http://schemas.openxmlformats.org/drawingml/2006/chart')
        root = ET.fromstring(xml_bytes)
        NS = 'http://schemas.openxmlformats.org/drawingml/2006/chart'
        plotArea = root.find('.//{%s}plotArea' % NS)
        if plotArea is None or root.find('.//{%s}dTable' % NS) is not None:
            return xml_bytes
        dTable = ET.SubElement(plotArea, '{%s}dTable' % NS)
        for tag in ['showHorzBorder', 'showVertBorder', 'showOutline', 'showKeys']:
            el = ET.SubElement(dTable, '{%s}%s' % (NS, tag))
            el.set('val', '1')
        # Remove the standalone legend — showKeys in dTable IS the legend row
        chart_el = root.find('{%s}chart' % NS)
        if chart_el is not None:
            leg = chart_el.find('{%s}legend' % NS)
            if leg is not None:
                chart_el.remove(leg)
        return ET.tostring(root, encoding='UTF-8', xml_declaration=True)

    with _zf.ZipFile(output_file, 'r') as _zin:
        _files = {n: _zin.read(n) for n in _zin.namelist()}
    for _cf in ['xl/charts/chart1.xml', 'xl/charts/chart3.xml']:
        if _cf in _files:
            _files[_cf] = _inject_dt(_files[_cf])
    import tempfile, shutil
    _tmp = output_file + '.tmp'
    with _zf.ZipFile(_tmp, 'w', _zf.ZIP_DEFLATED) as _zout:
        for _n, _d in _files.items():
            _zout.writestr(_n, _d)
    shutil.move(_tmp, output_file)
    print(f"Report saved to: {output_file}")

def has_ratings_sheet(filepath):
    try:
        import zipfile
        with zipfile.ZipFile(filepath, 'r') as z:
            wb_xml = z.read('xl/workbook.xml').decode('utf-8', errors='ignore')
            return 'ALL_SIPL_Ratings_List' in wb_xml
    except Exception:
        return False

def find_input_file():
    import glob
    if len(sys.argv) > 1:
        return sys.argv[1]
    # Only .xlsx files, exclude our own _REPORT outputs
    all_xlsx = glob.glob(os.path.join(SCRIPT_DIR, '*.xlsx'))
    all_xlsx = [f for f in all_xlsx if not os.path.basename(f).endswith('_REPORT.xlsx')]
    matches = all_xlsx  # accept any xlsx file
    if not matches:
        return None
    if len(matches) == 1:
        return matches[0]
    print("Multiple .xlsx files found. Please choose one:\n")
    for i, f in enumerate(matches, 1):
        print(f"  [{i}] {os.path.basename(f)}")
    print()
    while True:
        choice = input(f"Enter number (1-{len(matches)}): ").strip()
        if choice.isdigit() and 1 <= int(choice) <= len(matches):
            return matches[int(choice) - 1]
        print(f"  Invalid choice. Enter a number between 1 and {len(matches)}.")

def main():
    input_file = find_input_file()
    if not input_file:
        print("ERROR: No .xlsx file found. Place the SIPL xlsx file in the same folder as this script.")
        sys.exit(1)

    # Output filename mirrors input filename
    input_stem = os.path.splitext(os.path.basename(input_file))[0]
    output_file = os.path.join(SCRIPT_DIR, f'{input_stem}_REPORT.xlsx')

    # If file is already open/locked, try adding a timestamp
    import datetime
    if os.path.exists(output_file):
        try:
            os.rename(output_file, output_file)  # test if locked
        except PermissionError:
            ts = datetime.datetime.now().strftime('%H%M%S')
            output_file = os.path.join(SCRIPT_DIR, f'SIPL_Report_{ts}.xlsx')

    print(f"Reading file: {input_file}")
    xl = pd.ExcelFile(input_file)
    sheet = 'ALL_SIPL_Ratings_List' if 'ALL_SIPL_Ratings_List' in xl.sheet_names else xl.sheet_names[0]
    df = xl.parse(sheet)

    print("\n--- Running Validation Checks ---")
    issues = validate_file(df)

    if issues:
        print("\nFILE IS NOT CORRECT TO MAKE REPORT")
        print("\nIssues found:")
        for issue in issues:
            print(f"  - {issue}")
    else:
        print("\nFILE IS CORRECT TO MAKE REPORT")
        print("\nProceeding to generate report...")
        build_report(df, input_file, output_file)
        print("Report generated successfully!")

if __name__ == '__main__':
    main()
